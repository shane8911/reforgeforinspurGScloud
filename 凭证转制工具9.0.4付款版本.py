import tkinter as tk
from tkinter import filedialog
import pandas as pd
import uuid
import os
from datetime import datetime
from tkinter import messagebox
from tkinter import simpledialog
import ntplib
import sys
import qrcode

import pytz

from datetime import datetime, timezone

from PIL import Image, ImageTk




def convert_columns_to_string(df):
    for col in df.columns:
        df[col] = df[col].apply(lambda x: '' if pd.isna(x) else str(x))
    return df
def convert_accounting_direction(df):
    if '记账方向' in df.columns:
        df['记账方向'] = df['记账方向'].replace({'1': '借', '2': '贷'})
    return df


def create_target_format_df(target_columns, original_df, mapping):
    target_df = pd.DataFrame(columns=target_columns)
    for target_col, original_col in mapping.items():
        if original_col in original_df.columns:
            target_df[target_col] = original_df[original_col].astype(str)
    return target_df

def generate_uuid_mapping(original_ids):
    """为原始ID生成UUID映射"""
    return {oid: str(uuid.uuid4()) for oid in set(original_ids)}

def replace_ids_with_uuid(df, id_column, uuid_mapping):
    """使用UUID映射替换ID"""
    if id_column in df.columns:
        df[id_column] = df[id_column].apply(lambda x: uuid_mapping.get(x, x))
    return df

def load_files():
    global original_format_path, target_format_path
    original_format_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xls *.xlsx *.xlsm")])
    if original_format_path:  # 检查是否选择了文件
        process_data()
        messagebox.showinfo("加载成功", "文件加载成功！")
columns_mapping = {
    "凭证头": {
        "凭证ID": "凭证内码",
        "凭证编号": "凭证编号",
        "凭证日期": "凭证日期",
        "凭证类型编号": "凭证类型",
        "核算组织编号": "单位编号",
        "会计期间序号": "会计期间",
        "附件张数": "附件张数",
        "年度": "会计年度"
    },
    "凭证分录": {
        "凭证分录ID": "分录内码",
        "凭证ID": "凭证内码",
        "核算组织编号": "单位编号",
        "核算账簿编号": "单位编号",
        "摘要": "摘要",
        "凭证分录编号": "分录编号",
        "科目编号": "科目编号",
        "分录金额": "金额",
        "记账方向": "方向"
    },
    "凭证辅助": {
        "凭证分录ID": "分录内码",
        "凭证ID": "凭证内码",
        "核算组织编号": "单位编号",
        "核算账簿编号": "单位编号",
        "凭证辅助编号": "原始编号",
        "科目编号": "科目编号",
        "核算部门编号": "部门编号",
        "往来单位编号": "往来单位",
        "币种编号": "外币编号",
        "记账方向": "记账方向",
        "数量": "数量",
        "单价": "单价",
        "外币金额": "外币",
        "汇率": "汇率",
        "金额": "金额",
        "业务日期": "业务日期",
        "用途": "用途",
        "结算日期": "结算日期",
        "票据号": "票据号"
    }
}

# 目标格式的列名定义
target_voucher_head_columns = ['凭证ID', '凭证编号', '凭证日期', '凭证类型编号', '会计主管编号', '核算组织编号', 
                               '会计期间序号', '批准人编号', '审核人编号', '记账人编号', '出纳编号', '附件张数', '年度', 
                               '是否批准', '是否审核', '是否记账', '是否完整', '核算账簿编号', '是否作废', '制单人编号', 
                               '经办人编号', '密级ID', '密级名称', '是否签字']

target_voucher_entry_columns = ['凭证分录ID', '凭证ID', '核算组织编号', '核算账簿编号', '摘要', 
                                '凭证分录编号', '科目编号', '分录金额', '记账方向']

target_voucher_auxiliary_columns = ['凭证分录ID', '凭证ID', '核算组织编号', '核算账簿编号', '凭证辅助编号', '科目编号', 
                                    '核算部门编号', '往来单位编号', '核算人员编号', '币种编号', '记账方向', '数量', '单价', 
                                    '金额', '业务日期', '业务号', '经办人', 'SJ01', 'SJ02', 'SJ03', 
                                    'SJ04', 'SJ05', 'SM01', 'SM02', 'SM03', 'SM04', 'SM05', 'SM06', 'SM07', 'SM08', 
                                    'SM09', 'SM10', '结算方式', '结算号', '用途', '结算日期', '现金项目', '核算项目99', 
                                    '核算项目98', '核算项目97', '核算项目96', '核算项目95', '核算项目94', '核算项目93', 
                                    '核算项目92', '核算项目91', '核算项目90', '核算项目89', '核算项目88', '核算项目87', 
                                    '核算项目86', '核算项目85', '核算项目84', '核算项目83', '核算项目82', '核算项目81', 
                                    '核算项目80', '核算项目79', '核算项目78', '核算项目77', '核算项目76', '核算项目75', 
                                    '核算项目74', '核算项目73', '核算项目72', '核算项目71', '核算项目70', '核算项目69', 
                                    '核算项目68', '核算项目67', '核算项目66', '核算项目65', '核算项目64', '核算项目63', 
                                    '核算项目62', '核算项目61', '核算项目60', '核算项目59', '核算项目58', '核算项目57', 
                                    '核算项目56', '核算项目55', '核算项目54', '核算项目53', '核算项目52', '核算项目51', 
                                    '核算项目50', '核算项目49', '核算项目48', '核算项目47', '核算项目46', '核算项目45', 
                                    '核算项目44', '核算项目43', '核算项目42', '核算项目41', '核算项目40', '核算项目39', 
                                    '核算项目38', '核算项目37', '核算项目36', '核算项目35', '核算项目34', '核算项目33', 
                                    '核算项目32', '核算项目31', '核算项目30', '核算项目29', '核算项目28', '核算项目27', 
                                    '核算项目26', '核算项目25', '核算项目24', '核算项目23', '核算项目22', '核算项目21', 
                                    '核算项目20', '核算项目19', '核算项目18', '核算项目17', '核算项目16', '核算项目15', 
                                    '核算项目14', '核算项目13', '核算项目12', '核算项目11', '核算项目10', '核算项目09', 
                                    '核算项目08', '核算项目07', '核算项目06', '核算项目05', '核算项目04', '核算项目03', 
                                    '核算项目02', '核算项目01', '业务字典99', '业务字典98', '业务字典97', '业务字典96', 
                                    '业务字典95', '业务字典94', '业务字典93', '业务字典92', '业务字典91', '业务字典90', 
                                    '业务字典89', '业务字典88', '业务字典87', '业务字典86', '业务字典85', '业务字典84', 
                                    '业务字典83', '业务字典82', '业务字典81', '业务字典80', '业务字典79', '业务字典78', 
                                    '业务字典77', '业务字典76', '业务字典75', '业务字典74', '业务字典73', '业务字典72', 
                                    '业务字典71', '业务字典70', '业务字典69', '业务字典68', '业务字典67', '业务字典66', 
                                    '业务字典65', '业务字典64', '业务字典63', '业务字典62', '业务字典61', '业务字典60', 
                                    '业务字典59', '业务字典58', '业务字典57', '业务字典56', '业务字典55', '业务字典54', 
                                    '业务字典53', '业务字典52', '业务字典51', '业务字典50', '业务字典49', '业务字典48', 
                                    '业务字典47', '业务字典46', '业务字典45', '业务字典44', '业务字典43', '业务字典42', 
                                    '业务字典41', '业务字典40', '业务字典39', '业务字典38', '业务字典37', '业务字典36', 
                                    '业务字典35', '业务字典34', '业务字典33', '业务字典32', '业务字典31', '业务字典30', 
                                    '业务字典29', '业务字典28', '业务字典27', '业务字典26', '业务字典25', '业务字典24', 
                                    '业务字典23', '业务字典22', '业务字典21', '业务字典20', '业务字典19', '业务字典18', 
                                    '业务字典17', '业务字典16', '业务字典15', '业务字典14', '业务字典13', '业务字典12', 
                                    '业务字典11', '业务字典10', '业务字典09', '业务字典08', '业务字典07', '业务字典06', 
                                    '业务字典05', '业务字典04', '业务字典03', '业务字典02', '业务字典01', '票据号']
def process_data():
    original_sheets = pd.read_excel(original_format_path, sheet_name=None, dtype=str)

    voucher_id_mapping = generate_uuid_mapping(pd.concat([original_sheets['Sheet1']['凭证内码'], 
                                                         original_sheets['Sheet2']['凭证内码'], 
                                                         original_sheets['Sheet3']['凭证内码']]).unique())
    voucher_entry_id_mapping = generate_uuid_mapping(pd.concat([original_sheets['Sheet2']['分录内码'], 
                                                               original_sheets['Sheet3']['分录内码']]).unique())

    original_sheets['Sheet1'] = replace_ids_with_uuid(original_sheets['Sheet1'], '凭证内码', voucher_id_mapping)
    original_sheets['Sheet2'] = replace_ids_with_uuid(original_sheets['Sheet2'], '凭证内码', voucher_id_mapping)
    original_sheets['Sheet2'] = replace_ids_with_uuid(original_sheets['Sheet2'], '分录内码', voucher_entry_id_mapping)
    original_sheets['Sheet3'] = replace_ids_with_uuid(original_sheets['Sheet3'], '凭证内码', voucher_id_mapping)
    original_sheets['Sheet3'] = replace_ids_with_uuid(original_sheets['Sheet3'], '分录内码', voucher_entry_id_mapping)

    for sheet_name, df in original_sheets.items():
        original_sheets[sheet_name] = convert_columns_to_string(df)

    original_sheet2_df = convert_accounting_direction(original_sheets['Sheet2'])
    original_sheet3_df = convert_accounting_direction(original_sheets['Sheet3'])
  
    global converted_voucher_head_with_blanks, converted_voucher_entry_with_blanks, converted_voucher_auxiliary_with_blanks
    # 根据 '项目编号' 映射到 '核算项目01' 和 '核算项目04'
    original_sheet3_df['核算项目01'] = original_sheet3_df['项目编号'].apply(lambda x: '010010001' if x == '0003' else ('010020001' if x == '000h' else ''))
    original_sheet3_df['核算项目04'] = original_sheet3_df['项目编号'].apply(lambda x: x if x not in ['0003', '000h'] else '')

    original_sheet3_df = convert_accounting_direction(original_sheet3_df)

    # 确保目标格式中的 '核算项目01' 和 '核算项目04' 被正确映射
    target_voucher_auxiliary_columns.extend(['核算项目01', '核算项目04'])
    columns_mapping["凭证辅助"].update({'核算项目01': '核算项目01', '核算项目04': '核算项目04'})

    converted_voucher_auxiliary_with_blanks = create_target_format_df(target_voucher_auxiliary_columns, original_sheet3_df, columns_mapping["凭证辅助"])


    

    converted_voucher_head_with_blanks = create_target_format_df(target_voucher_head_columns, original_sheets['Sheet1'], columns_mapping["凭证头"])
    converted_voucher_entry_with_blanks = create_target_format_df(target_voucher_entry_columns, original_sheet2_df, columns_mapping["凭证分录"])
    converted_voucher_auxiliary_with_blanks = create_target_format_df(target_voucher_auxiliary_columns, original_sheet3_df, columns_mapping["凭证辅助"])
def update_voucher_number():
    user_input = simpledialog.askstring("输入凭证编号", "请输入凭证编号（4位数字，如0003）:")
    if user_input and user_input.isdigit() and len(user_input) == 4:
        update_voucher_number_in_dataframe(user_input)
    else:
        messagebox.showwarning("输入错误", "凭证编号必须是4位数字。")

def update_voucher_number_in_dataframe(start_number):
    start_num = int(start_number)
    for i in range(len(converted_voucher_head_with_blanks)):
        converted_voucher_head_with_blanks.at[i, '凭证编号'] = f"记{str(start_num).zfill(4)}"
        start_num += 1
import tkinter as tk
from tkinter import filedialog
import openpyxl

def fill_signer_id(excel_file_path):
    映射 = {
        "09": "杨石瑀",
        "03": "liuchuyi",
        "02": "liuchuyi",
        "10": "suyu",
        "40": "suyu",
        "06": "杨石瑀",
        "07": "翁兴传",
        "04": "翁兴传",
        "2401": "翁兴传",
        "2501": "翁兴传",
        "2601": "翁兴传",
        "2701": "翁兴传4"
    }
    
    # 打开Excel文件
    workbook = openpyxl.load_workbook(excel_file_path)
    源工作表 = workbook["凭证辅助"]
    目标工作表 = workbook["凭证头"]
    
    源行数 = 源工作表.max_row
    
    
    制单人编号列表 = []
    上方单元格为空 = True  

    for i in range(2, 源行数 + 1):
        核算部门编号 = 源工作表.cell(row=i, column=7).value
        上方核算部门编号 = 源工作表.cell(row=i-1, column=7).value if i > 2 else None  
        
        if 核算部门编号 is not None and (上方核算部门编号 is None or 上方单元格为空):
            编号 = str(核算部门编号)
            if 编号 in 映射:
                制单人编号列表.append(映射[编号])
        
        上方单元格为空 = (核算部门编号 is None)  
    
    
    for row in range(2, 目标工作表.max_row + 1):
        目标工作表.cell(row=row, column=20).value = None
    
   
    for i, value in enumerate(制单人编号列表, start=2):
        目标工作表.cell(row=i, column=20).value = value
    
   
    workbook.save(excel_file_path)
    print("制单人编号填写完成。")
    messagebox.showinfo("提示", "制单人编号填写完成。")



def save_output():
    global converted_voucher_auxiliary_with_blanks

    output_path = filedialog.asksaveasfilename(defaultextension=".xlsx")
    converted_voucher_entry_with_blanks.iloc[:, 6] = pd.to_numeric(converted_voucher_entry_with_blanks.iloc[:, 6], errors='coerce').fillna(0).round(2)
    
    if output_path:  
        

       
        dept_code_mapping = {
            '2401': ('22410302', '04'),
            '2601': ('22410301', '04'),
            '220411': ('220411', '04')
        }
        
        # 遍历每一行处理部门编号映射
        for index in converted_voucher_auxiliary_with_blanks.index:
            dept_code = converted_voucher_auxiliary_with_blanks.loc[index, '核算部门编号']
            if str(dept_code) in dept_code_mapping:
                new_subject_code, new_dept_code = dept_code_mapping[str(dept_code)]
                # 更新凭证辅助表
                converted_voucher_auxiliary_with_blanks.loc[index, '科目编号'] = new_subject_code
                converted_voucher_auxiliary_with_blanks.loc[index, '核算部门编号'] = new_dept_code
                # 更新凭证分录表中对应的行
                entry_index = converted_voucher_entry_with_blanks.index[converted_voucher_entry_with_blanks['凭证分录ID'] == converted_voucher_auxiliary_with_blanks.loc[index, '凭证分录ID']]
                if len(entry_index) > 0:
                    converted_voucher_entry_with_blanks.loc[entry_index, '科目编号'] = new_subject_code

        # 处理科目编号 22030202 -> 220401 的替换
        # 更新凭证辅助表
        converted_voucher_auxiliary_with_blanks.loc[converted_voucher_auxiliary_with_blanks['科目编号'] == '22030202', '科目编号'] = '220401'
        # 更新凭证分录表
        converted_voucher_entry_with_blanks.loc[converted_voucher_entry_with_blanks['科目编号'] == '22030202', '科目编号'] = '220401'
        # 硬编码校验：如果凭证分录表第二行的科目编号是 '11230202'，则替换为 '112302'
        # 更新凭证分录表
        converted_voucher_entry_with_blanks.loc[converted_voucher_entry_with_blanks['科目编号'] == '11230202', '科目编号'] = '112302'

        


       
        converted_voucher_auxiliary_with_blanks.loc[converted_voucher_auxiliary_with_blanks['科目编号'] == '11230202', '科目编号'] = '112302'

        converted_voucher_entry_with_blanks.loc[converted_voucher_entry_with_blanks['科目编号'] == '11230202', '科目编号'] = '112302'
       
        mask_10021401_aux = converted_voucher_auxiliary_with_blanks['科目编号'] == '10021401'
        converted_voucher_auxiliary_with_blanks.loc[mask_10021401_aux, '业务字典01'] = '87205022200201018726'
        converted_voucher_auxiliary_with_blanks.loc[mask_10021401_aux, '科目编号'] = '100202'
        
        
        affected_entry_ids = converted_voucher_auxiliary_with_blanks.loc[mask_10021401_aux, '凭证分录ID'].tolist()
        
        converted_voucher_entry_with_blanks.loc[converted_voucher_entry_with_blanks['凭证分录ID'].isin(affected_entry_ids), '科目编号'] = '100202'

        
        mask_10020201_aux = converted_voucher_auxiliary_with_blanks['科目编号'] == '10020201'
        converted_voucher_auxiliary_with_blanks.loc[mask_10020201_aux, '业务字典01'] = '318166008853'
        converted_voucher_auxiliary_with_blanks.loc[mask_10020201_aux, '科目编号'] = '100202'
        
        
        affected_entry_ids = converted_voucher_auxiliary_with_blanks.loc[mask_10020201_aux, '凭证分录ID'].tolist()
        
        converted_voucher_entry_with_blanks.loc[converted_voucher_entry_with_blanks['凭证分录ID'].isin(affected_entry_ids), '科目编号'] = '100202'

        
        mask_10020101_aux = converted_voucher_auxiliary_with_blanks['科目编号'] == '10020101'
        converted_voucher_auxiliary_with_blanks.loc[mask_10020101_aux, '业务字典01'] = '0200003419201092331'
        converted_voucher_auxiliary_with_blanks.loc[mask_10020101_aux, '科目编号'] = '100202'
        
        
        affected_entry_ids = converted_voucher_auxiliary_with_blanks.loc[mask_10020101_aux, '凭证分录ID'].tolist()

        converted_voucher_entry_with_blanks.loc[converted_voucher_entry_with_blanks['凭证分录ID'].isin(affected_entry_ids), '科目编号'] = '100202'

       
        mask_10020301_aux = converted_voucher_auxiliary_with_blanks['科目编号'] == '10020301'
        converted_voucher_auxiliary_with_blanks.loc[mask_10020301_aux, '业务字典01'] = '11001046600053003418'
        converted_voucher_auxiliary_with_blanks.loc[mask_10020301_aux, '科目编号'] = '100202'
        
        
        affected_entry_ids = converted_voucher_auxiliary_with_blanks.loc[mask_10020301_aux, '凭证分录ID'].tolist()
       
        converted_voucher_entry_with_blanks.loc[converted_voucher_entry_with_blanks['凭证分录ID'].isin(affected_entry_ids), '科目编号'] = '100202'

       
        mask_10020401_aux = converted_voucher_auxiliary_with_blanks['科目编号'] == '10020401'
        converted_voucher_auxiliary_with_blanks.loc[mask_10020401_aux, '业务字典01'] = '042701040003025'
        converted_voucher_auxiliary_with_blanks.loc[mask_10020401_aux, '科目编号'] = '100202'
        
       
        affected_entry_ids = converted_voucher_auxiliary_with_blanks.loc[mask_10020401_aux, '凭证分录ID'].tolist()
        
        converted_voucher_entry_with_blanks.loc[converted_voucher_entry_with_blanks['凭证分录ID'].isin(affected_entry_ids), '科目编号'] = '100202'

        
        mask_10020402_aux = converted_voucher_auxiliary_with_blanks['科目编号'] == '10020402'
        converted_voucher_auxiliary_with_blanks.loc[mask_10020402_aux, '业务字典01'] = '021401040004587'
        converted_voucher_auxiliary_with_blanks.loc[mask_10020402_aux, '科目编号'] = '100202'
        
        
        affected_entry_ids = converted_voucher_auxiliary_with_blanks.loc[mask_10020402_aux, '凭证分录ID'].tolist()
        
        converted_voucher_entry_with_blanks.loc[converted_voucher_entry_with_blanks['凭证分录ID'].isin(affected_entry_ids), '科目编号'] = '100202'

        
        mask_10020701_aux = converted_voucher_auxiliary_with_blanks['科目编号'] == '10020701'
        converted_voucher_auxiliary_with_blanks.loc[mask_10020701_aux, '业务字典01'] = '7115010182600002294'
        converted_voucher_auxiliary_with_blanks.loc[mask_10020701_aux, '科目编号'] = '100202'
        
        
        affected_entry_ids = converted_voucher_auxiliary_with_blanks.loc[mask_10020701_aux, '凭证分录ID'].tolist()
        
        converted_voucher_entry_with_blanks.loc[converted_voucher_entry_with_blanks['凭证分录ID'].isin(affected_entry_ids), '科目编号'] = '100202'

        
        mask_10021002_aux = converted_voucher_auxiliary_with_blanks['科目编号'] == '10021002'
        converted_voucher_auxiliary_with_blanks.loc[mask_10021002_aux, '业务字典01'] = '137411512010000405'
        converted_voucher_auxiliary_with_blanks.loc[mask_10021002_aux, '科目编号'] = '100202'
        
        
        affected_entry_ids = converted_voucher_auxiliary_with_blanks.loc[mask_10021002_aux, '凭证分录ID'].tolist()
        
        converted_voucher_entry_with_blanks.loc[converted_voucher_entry_with_blanks['凭证分录ID'].isin(affected_entry_ids), '科目编号'] = '100202'

        
        mask_10021101_aux = converted_voucher_auxiliary_with_blanks['科目编号'] == '10021101'
        converted_voucher_auxiliary_with_blanks.loc[mask_10021101_aux, '业务字典01'] = '11006071193403'
        converted_voucher_auxiliary_with_blanks.loc[mask_10021101_aux, '科目编号'] = '100202'
        
        
        affected_entry_ids = converted_voucher_auxiliary_with_blanks.loc[mask_10021101_aux, '凭证分录ID'].tolist()
        
        converted_voucher_entry_with_blanks.loc[converted_voucher_entry_with_blanks['凭证分录ID'].isin(affected_entry_ids), '科目编号'] = '100202'

        
        converted_voucher_entry_with_blanks.iloc[1:, 6] = converted_voucher_auxiliary_with_blanks.iloc[1:, 5]
        
        
        current_group_value = None
        group_first_value = None
        
        
        for i in range(1, len(converted_voucher_auxiliary_with_blanks)):
            
            current_value = converted_voucher_auxiliary_with_blanks.iloc[i, 6]
            
            
            if pd.notna(current_value) and str(current_value).strip() != '':
                
                if current_group_value is None:
                    current_group_value = current_value
                    
                    group_first_value = converted_voucher_auxiliary_with_blanks.iloc[i, 7]
                
                
                converted_voucher_auxiliary_with_blanks.iloc[i, 7] = group_first_value
            else:
                
                current_group_value = None
                group_first_value = None

       
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            converted_voucher_head_with_blanks.to_excel(writer, sheet_name='凭证头', index=False)
            converted_voucher_entry_with_blanks.to_excel(writer, sheet_name='凭证分录', index=False)
            converted_voucher_auxiliary_with_blanks.to_excel(writer, sheet_name='凭证辅助', index=False)
    
        
        bill_number_col_index = converted_voucher_auxiliary_with_blanks.columns.get_loc("票据号")
        converted_voucher_entry_with_blanks.iloc[0, 6] = '112302'
        
        converted_voucher_auxiliary_with_blanks = converted_voucher_auxiliary_with_blanks.iloc[:, :(bill_number_col_index+1)]
       
        
        converted_voucher_auxiliary_with_blanks['金额'] = pd.to_numeric(
            converted_voucher_auxiliary_with_blanks['金额'], errors='coerce').fillna(0).round(2)

       
        if '单价' in converted_voucher_auxiliary_with_blanks.columns and '数量' in converted_voucher_auxiliary_with_blanks.columns:
            
            converted_voucher_auxiliary_with_blanks['数量'] = pd.to_numeric(converted_voucher_auxiliary_with_blanks['数量'], errors='coerce').fillna(0)

           
            converted_voucher_auxiliary_with_blanks['单价'] = converted_voucher_auxiliary_with_blanks.apply(
                lambda row: row['金额'] / row['数量'] if row['数量'] != 0 else '',
                axis=1
            )
        
            
        
        with pd.ExcelWriter(output_path, mode='a', engine='openpyxl', if_sheet_exists='replace') as writer:
            converted_voucher_auxiliary_with_blanks.to_excel(writer, sheet_name='凭证辅助', index=False)

        messagebox.showinfo("保存成功", f"结果已保存到：{output_path}")


import tkinter as tk
from tkinter import filedialog
import openpyxl




def select_file():
    file_path = filedialog.askopenfilename(title="选择待处理的Excel文件",
                                           filetypes=(("Excel files", "*.xlsx"), ("All files", "*.*")))
    if file_path:
        fill_signer_id(file_path)



root = tk.Tk()
root.title("数据转换工具")



# 设置窗口大小为384x384
root.geometry("384x384")

# 获取当前脚本所在的目录
current_directory = os.path.dirname(os.path.realpath(__file__))

# 构建背景图片的完整路径
bg_image_path = 'background.png'

# 加载背景图片
bg_image = Image.open(bg_image_path)
bg_photo = ImageTk.PhotoImage(bg_image)

# 创建一个标签来显示背景图片
bg_label = tk.Label(root, image=bg_photo)
bg_label.place(relwidth=1, relheight=1)

# 创建按钮 - 示例
load_button = tk.Button(root, text="加载文件", command=load_files)
load_button.place(x=20, y=330)  # 设置按钮位置

voucher_number_button = tk.Button(root, text="凭证编号", command=update_voucher_number)
voucher_number_button.place(x=90, y=330)  # 设置按钮位置

save_button = tk.Button(root, text="输出结果", command=save_output)
save_button.place(x=160, y=330)  # 设置按钮位置

pro_button = tk.Button(root, text="添加制单人", command=select_file)
pro_button.place(x=230, y=330)  # 设置按钮位置


# 运行主循环
root.mainloop()
